import openpyxl
from openpyxl.styles import Font
import io
import math
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')
from matplotlib.patches import Circle
from io import BytesIO
import logging
import json
import base64
import numpy as np
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.core.files.base import ContentFile
from django.http import HttpResponse, JsonResponse
from .models import Project, Attachment, Profile, Protocol
from .forms import ProjectForm, AttachmentForm, ProfileForm, ProtocolForm
from openpyxl import Workbook
from django.core.paginator import Paginator

def index(request):
    # Получаем данные с пагинацией
    projects = Paginator(Project.objects.all().order_by('-id'), 10)
    attachments = Paginator(Attachment.objects.all().order_by('-id'), 10)
    profiles = Paginator(Profile.objects.all().order_by('-created_at'), 10)
    protocols = Paginator(Protocol.objects.all().order_by('-created_at'), 10)

    # Получаем номер страницы из GET-параметра
    page_number = request.GET.get('page')

    return render(request, 'index.html', {
        'projects': projects.get_page(page_number),
        'attachments': attachments.get_page(page_number),
        'profiles': profiles.get_page(page_number),
        'protocols': protocols.get_page(page_number)
    })

def profil_page(request):
    return render(request, 'profil.html')

#для строки поиска
def search_view(request):
    query=request.GET.get('q')
    return render(request, 'search_results.html', {'query': query})

@csrf_exempt
def generate_profile(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            #Валидация данных
            if not data.get('title') or not data.get('ground_lengths'):
                return JsonResponse({'error': 'Недостаточно данных'}, status=400)
            # Здесь может быть ваша логика обработки
            response_data = {
                'status': 'success',
                'title': data['title'],
                'message': 'Профиль успешно сгенерирован'
            }
            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Метод не разрешен'}, status=405)

logger=logging.getLogger(__name__)


@csrf_exempt
def generate_profile_image(request):
    if request.method == 'POST':
        print("\n=== НОВЫЙ ЗАПРОС ===")  # Отладочный разделитель

        # 1. Логирование сырых данных
        raw_data = request.body.decode('utf-8')
        print("Сырые данные запроса:", raw_data)

        try:
            # 2. Парсинг JSON
            data = json.loads(raw_data)
            print("Распарсенные данные:", json.dumps(data, indent=2))

            # 3. Валидация обязательных полей
            required_fields = ['title', 'ground_lengths', 'ground_heights']
            missing_fields = [field for field in required_fields if field not in data]
            if missing_fields:
                error_msg = f"Отсутствуют обязательные поля: {', '.join(missing_fields)}"
                print(error_msg)
                return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

            # 4. Обработка данных грунта
            try:
                ground_lengths = [float(x.strip()) for x in data['ground_lengths'].split(',')]
                ground_heights = [float(x.strip()) for x in data['ground_heights'].split(',')]
            except ValueError as e:
                error_msg = f"Ошибка преобразования данных грунта: {str(e)}"
                print(error_msg)
                return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

            # 5. Проверка согласованности данных
            if len(ground_lengths) != len(ground_heights):
                error_msg = f"Несоответствие количества точек: {len(ground_lengths)} длин vs {len(ground_heights)} высот"
                print(error_msg)
                return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

            if len(ground_lengths) < 2:
                error_msg = "Недостаточно точек для построения графика (минимум 2)"
                print(error_msg)
                return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

            # 6. Обработка труб (если есть)
            pipes_data = []
            for pipe in data.get('pipes', []):
                try:
                    lengths = [float(x.strip()) for x in pipe['lengths'].split(',')]
                    heights = [float(x.strip()) for x in pipe['heights'].split(',')]
                    pipes_data.append({
                        'name': pipe.get('name', 'Труба'),
                        'lengths': lengths,
                        'heights': heights
                    })
                except Exception as e:
                    print(f"Ошибка обработки трубы: {str(e)}")
                    continue

            # 7. Создание графика
            plt.figure(figsize=(14, 8))

            # График грунта
            plt.plot(ground_lengths, ground_heights, 'b-', label='Уровень грунта', linewidth=2)

            # Графики труб
            for pipe in pipes_data:
                plt.plot(pipe['lengths'], pipe['heights'], '-', linewidth=3, label=pipe['name'])

            # Настройки графика
            plt.title(data['title'], fontsize=14, pad=20)
            plt.xlabel('Длина прокола, м', fontsize=12)
            plt.ylabel('Отметки высот, м', fontsize=12)
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.legend()

            # 8. Сохранение в буфер
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()

            # 9. Подготовка ответа
            buffer.seek(0)
            print("График успешно сгенерирован")
            return HttpResponse(buffer.getvalue(), content_type='image/png')

        except json.JSONDecodeError as e:
            error_msg = f"Ошибка декодирования JSON: {str(e)}"
            print(error_msg)
            return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

        except Exception as e:
            error_msg = f"Непредвиденная ошибка: {str(e)}"
            print(error_msg)
            return JsonResponse({'status': 'error', 'message': error_msg}, status=500)

    error_msg = "Метод не разрешен (требуется POST)"
    print(error_msg)
    return JsonResponse({'status': 'error', 'message': error_msg}, status=405)

@csrf_exempt
def test_matplotlib(request):
    try:
        plt.figure()
        plt.plot([0, 1, 2], [0, 1, 0])
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        plt.close()
        buffer.seek(0)
        return HttpResponse(buffer, content_type='image/png')
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def save_profile(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Проверяем обязательные поля
            if not data.get('address') or not data.get('image_data'):
                return JsonResponse({'status': 'error', 'message': 'Адрес и изображение обязательны'}, status=400)

            # Декодируем изображение
            format, imgstr = data['image_data'].split(';base64,')
            ext = format.split('/')[-1]
            image_file = ContentFile(
                base64.b64decode(imgstr),
                name=f"profile_{data['address'][:50]}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
            )

            # Создаем профиль
            profile = Profile(
                address=data['address'],
                description=data.get('description', ''),
                image=image_file
            )
            profile.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Профиль сохранен',
                'profile_id': profile.id,
                'address': profile.address,
                'created_at': profile.created_at.strftime("%d.%m.%Y %H:%M")
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Метод не разрешен'}, status=405)

@csrf_exempt
def generate_protocol(request):
    if request.method == 'POST':
        address = request.POST.get('address')
        length = float(request.POST.get('length'))
        rod_length = float(request.POST.get('rod_length'))
        head_level = float(request.POST.get('head_level'))
        depth = float(request.POST.get('depth'))
        system = request.POST.get('system')
        location = request.POST.get('location')
        supervisor = request.POST.get('supervisor')
        engineer = request.POST.get('engineer')
        expansion = request.POST.get('expansion')

        wb = Workbook()
        ws = wb.active
        ws.title = "Протокол бурения"

        ws.append(['ПРОТОКОЛ БУРЕНИЯ'])
        ws.append(['Адрес', address])
        ws.append(['Буровая система', system])
        ws.append(['Длина прокола (м)', length])
        ws.append(['Расширения (мм)', expansion])
        ws.append(['Система локации', location])
        ws.append(['Руководитель бурения', supervisor])
        ws.append([])

        ws.append(['№ штанги', 'Длина пилотной скважины, м', 'Угол, %', 'Глубина, м'])

        num_rods = math.ceil(length / rod_length)
        for i in range(1, num_rods + 1):
            rod_len = min(i * rod_length, length)
            hypotenuse = rod_len
            cos_angle = rod_length / hypotenuse
            angle_percent = round((1 - cos_angle) * 100, 2)
            depth_head = round(head_level - (depth / length * rod_len), 2)
            ws.append([i, rod_len, angle_percent, depth_head])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        protocol = Protocol(address=address, description=f"Протокол по адресу {address}")
        protocol.file.save(f"protocol_{address}.xlsx", ContentFile(buffer.getvalue()))
        protocol.save()
    return redirect('index')

def protocol(request):
    if request.method == 'POST':
        form = ProtocolForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data

            # Список глубин
            ground_levels = list(map(float, data['ground_levels'].split(',')))
            pipe_depths = list(map(float, data['pipe_depths'].split(',')))
            expansions = [e.strip() for e in data['expansions'].split(',')]

            rod_length = data['rod_length']
            total_length = data['total_length']

            # Количество штанг (с округлением вверх)
            rod_count = math.ceil(total_length / rod_length)

            # Расчёт данных по каждой штанге
            protocol_data = []
            for i in range(rod_count):
                length = round((i + 1) * rod_length, 2)

                if i < len(ground_levels) and i < len(pipe_depths):
                    depth = round(ground_levels[i] - pipe_depths[i], 2)
                else:
                    depth = 0

                # гипотенуза = длина штанги (род)
                hypotenuse = math.sqrt(rod_length**2 + depth**2) if depth else rod_length
                angle_percent = round((depth / hypotenuse) * 100, 2) if hypotenuse else 0

                protocol_data.append({
                    'rod_number': i + 1,
                    'length': length,
                    'angle': angle_percent,
                    'depth': depth,
                })

            context = {
                'form': form,
                'address': data['address'],
                'drill_system': data['drill_system'],
                'location_system': data['location_system'],
                'drilling_supervisor': data['drilling_supervisor'],
                'construction_supervisor': data['construction_supervisor'],
                'total_length': total_length,
                'rod_length': rod_length,
                'expansions': expansions,
                'protocol_data': protocol_data,
                'submitted': True
            }
            return render(request, 'protocol.html', context)

    else:
        form = ProtocolForm()

    return render(request, 'protocol.html', {'form': form})

#КопироватьРедактировать
def export_protocol_excel(request):
    if request.method == 'POST':
        rod_length = float(request.POST.get('rod_length'))
        total_length = float(request.POST.get('total_length'))
        ground_levels = list(map(float, request.POST.get('ground_levels').split(',')))
        pipe_depths = list(map(float, request.POST.get('pipe_depths').split(',')))

        rod_count = math.ceil(total_length / rod_length)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Протокол"

        # Заголовки
        headers = ['№ штанги', 'Длина пилотной скважины (м)', 'Угол наклона (%)', 'Глубина головки (м)']
        ws.append(headers)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in col:
                cell.font = Font(bold=True)

        for i in range(rod_count):
            length = round((i + 1) * rod_length, 2)
            if i < len(ground_levels) and i < len(pipe_depths):
                depth = round(ground_levels[i] - pipe_depths[i], 2)
            else:
                depth = 0

            hypotenuse = math.sqrt(rod_length**2 + depth**2) if depth else rod_length
            angle_percent = round((depth / hypotenuse) * 100, 2) if hypotenuse else 0

            ws.append([i + 1, length, angle_percent, depth])

        # Сохраняем в поток
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=protocol.xlsx'
        return response

    return HttpResponse("Ошибка: только POST запрос", status=400)

def protocol_view(request):
    if request.method == 'POST':
        form = ProtocolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = ProtocolForm()
    return render(request, 'protocol.html', {'form': form})


def save_profile_document(request):
    return HttpResponse(" Заглушка для save_profile_document - функция работает")